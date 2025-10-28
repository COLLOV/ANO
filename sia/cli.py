from __future__ import annotations

import argparse
import csv
import json
import logging
import sys
from pathlib import Path
import os
from concurrent.futures import ThreadPoolExecutor
from typing import Iterable, List, Optional, Dict, Any
from collections import Counter

import orjson
import tomli
from tqdm import tqdm

from .config import load_settings
from .llm_client import LLMClient
from .categorizer import categorize_text
from .taxonomy import Taxonomy, LLMTaxonomyConsolidator


def setup_logging(level: str) -> None:
    logging.basicConfig(
        level=level,
        format="%(asctime)s %(levelname)s %(name)s - %(message)s",
    )


def read_csv_rows(path: Path) -> Iterable[dict]:
    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            yield row


def read_xlsx_rows(path: Path) -> Iterable[dict]:
    """Lire les lignes d'un fichier XLSX en dict.

    - Utilise la première feuille
    - La première ligne est l'en-tête
    - Pas de mécanisme de fallback: en cas d'erreur on laisse remonter
    """
    from openpyxl import load_workbook  # dépendance ajoutée via uv

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return

    headers = [
        (str(h).strip() if h is not None else "")
        for h in header
    ]
    logging.debug("XLSX headers: %s", headers)

    for row in rows_iter:
        # Mappe chaque cellule à son en-tête; si vide, génère un nom de colonne
        out: dict = {}
        for i, val in enumerate(row):
            key = headers[i] if i < len(headers) and headers[i] else f"col_{i+1}"
            out[key] = val
        yield out


def read_rows(path: Path) -> Iterable[dict]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        logging.info("Lecture XLSX: %s", path)
        return read_xlsx_rows(path)
    elif suffix == ".csv" or not suffix:
        # Par défaut on considère CSV si pas d'extension
        logging.info("Lecture CSV: %s", path)
        return read_csv_rows(path)
    else:
        raise ValueError(f"Extension non supportée: {suffix}")


def _cli_provided(argv: List[str], opt: str) -> bool:
    """Retourne True si l'option CLI `opt` (ex: --input) est fournie explicitement."""
    for i, tok in enumerate(argv):
        if tok == opt:
            return True
        if tok.startswith(opt + "="):
            return True
    return False


def load_cli_config(cfg_path: Path, profile: Optional[str]) -> tuple[Dict[str, Any], Dict[str, str]]:
    """Charge un fichier TOML et retourne (options, env_overrides).

    Structure minimale supportée:
      # top-level
      input = "..."; output = "..."; format = "jsonl|csv"; consolidate = true|false
      save_intermediate = "..."; resume_consolidate_from = "..."
      workers = 20; limit = 0; consolidation_batch_size = 500
      [env]  # optionnel
      LLM_MODE = "api"; CLI_WORKERS = "20"; API_WORKERS = "1"

      # OU via profils
      [profiles.default]
      input = "..."
      [run]
      profile = "default"
    """
    with cfg_path.open("rb") as f:
        data = tomli.loads(f.read().decode("utf-8"))
    env_over = data.get("env") or {}
    opts: Dict[str, Any] = {}

    prof = profile or (data.get("run") or {}).get("profile")
    if prof:
        profiles = data.get("profiles") or {}
        if prof not in profiles:
            raise ValueError(f"Profil introuvable dans {cfg_path}: {prof}")
        opts = profiles[prof] or {}
    else:
        # top-level options
        opts = {k: v for k, v in data.items() if k not in {"env", "run", "profiles"}}
    return opts, {str(k): str(v) for k, v in env_over.items()}


def extract_text(row: dict) -> Optional[str]:
    for key in ("feedback", "text", "message", "comment"):
        if key in row and row[key]:
            return str(row[key])
    # Jira FR style
    if row.get("resume") or row.get("description"):
        return (row.get("resume", "").strip() + ". " + row.get("description", "").strip()).strip()
    return None


def main(argv: List[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Catégorisation de feedbacks via LLM")
    parser.add_argument(
        "--config",
        type=Path,
        default=None,
        help="Fichier de configuration TOML (défaut: ./sia.toml si présent)",
    )
    parser.add_argument("--profile", type=str, default=None, help="Nom de profil dans le fichier de config")
    parser.add_argument("--input", type=Path, default=Path("data/tickets_jira.csv"))
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--limit", type=int, default=0, help="Limiter le nombre de lignes (0 = tout)")
    # Si non fourni, on prendra la valeur depuis .env (CLI_WORKERS)
    parser.add_argument("--workers", type=int, default=None, help="Nombre de workers en parallèle (défaut: CLI_WORKERS ou 20)")
    parser.add_argument("--consolidate", action="store_true", help="Consolider les catégories/sous-catégories via LLM (2 passes)")
    parser.add_argument("--format", choices=["jsonl", "csv"], default="csv", help="Format de sortie (défaut: csv)")
    parser.add_argument("--save-intermediate", type=Path, default=None, help="Sauver les résultats intermédiaires (avant consolidation) en JSONL")
    parser.add_argument("--resume-consolidate-from", type=Path, default=None, help="Reprendre la consolidation depuis un JSONL intermédiaire")
    parser.add_argument("--consolidation-batch-size", type=int, default=None, help="Taille des lots LLM pour la consolidation (défaut: CONSOLIDATION_BATCH_SIZE)")
    parser.add_argument("--consolidation-rounds", type=int, default=None, help="Nombre de passes de consolidation (défaut: CONSOLIDATION_ROUNDS=1)")
    parser.add_argument("--max-categories", type=int, default=None, help="Nombre maximum de catégories cibles (fusion heuristique préalable)")
    parser.add_argument("--merge-threshold", type=float, default=None, help="Seuil de similarité (0-1) pour la fusion heuristique préalable")
    args = parser.parse_args(argv)

    # Gestion config: charger tôt pour injecter des overrides d'env AVANT load_settings()
    orig_argv = list(argv) if argv is not None else sys.argv[1:]
    cfg_opts: Dict[str, Any] = {}
    cfg_env: Dict[str, str] = {}
    cfg_path: Optional[Path] = None
    if args.config:
        cfg_path = args.config
    else:
        # Auto-charge le fichier local `sia.toml` s'il existe
        default_cfg = Path("sia.toml")
        if default_cfg.exists():
            cfg_path = default_cfg
    if cfg_path:
        cfg_opts, cfg_env = load_cli_config(cfg_path, args.profile)
        for k, v in cfg_env.items():
            os.environ[k] = v

    settings = load_settings()
    setup_logging(settings.log_level)
    client = LLMClient(settings)
    taxo = Taxonomy()
    consolidator = LLMTaxonomyConsolidator(client) if args.consolidate else None

    if args.resume_consolidate_from and args.limit:
        logging.warning("--limit est ignoré en mode reprise de consolidation")

    # Fusion config -> args si l'option n'est pas fournie en CLI
    def choose(opt: str, current, conv=lambda x: x):
        if _cli_provided(orig_argv, f"--{opt.replace('_','-')}"):
            return current
        if opt in cfg_opts and cfg_opts[opt] is not None:
            return conv(cfg_opts[opt])
        return current

    args.input = choose("input", args.input, Path)
    args.output = choose("output", args.output, lambda x: Path(x) if x else None)
    args.format = choose("format", args.format, str)
    args.limit = choose("limit", args.limit, int)
    args.workers = choose("workers", args.workers, lambda x: int(x) if x is not None else None)
    args.consolidate = choose("consolidate", args.consolidate, bool)
    args.save_intermediate = choose("save_intermediate", args.save_intermediate, lambda x: Path(x) if x else None)
    args.resume_consolidate_from = choose("resume_consolidate_from", args.resume_consolidate_from, lambda x: Path(x) if x else None)
    args.consolidation_batch_size = choose("consolidation_batch_size", args.consolidation_batch_size, int)
    args.consolidation_rounds = choose("consolidation_rounds", args.consolidation_rounds, int)
    args.max_categories = choose("max_categories", args.max_categories, lambda x: int(x) if x is not None else None)
    args.merge_threshold = choose("merge_threshold", args.merge_threshold, float)

    if not args.input.exists() and not args.resume_consolidate_from:
        logging.error("Input file not found: %s", args.input)
        return 2

    # Préparation de la sortie
    out_f = None
    csv_writer = None
    if args.format == "csv" and args.output:
        out_f = args.output.open("w", newline="", encoding="utf-8")
    elif args.format == "jsonl" and args.output:
        out_f = args.output.open("wb")

    count = 0
    buffered: List[dict] = []
    all_cats: set[str] = set()
    all_subs: set[str] = set()
    cat_freq: Counter[str] = Counter()
    input_fieldnames: List[str] = []

    # Pré-lire les lignes (jusqu'à --limit) pour pouvoir paralléliser proprement
    rows: List[dict] = []
    if not args.resume_consolidate_from:
        for row in read_rows(args.input):
            rows.append(row)
            if args.limit and len(rows) >= args.limit:
                break

    if rows and args.format == "csv":
        # Construire des en-têtes stables: ordre = 1er row puis ajout des nouvelles colonnes rencontrées
        base = list(rows[0].keys())
        seen = set(base)
        extras_ordered: list[str] = []
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    seen.add(k)
                    extras_ordered.append(k)
        input_fieldnames = base + extras_ordered
        output_fieldnames = input_fieldnames + [
            "category", "subcategories", "sentiment",
            "emotional_tone", "summary", "estimated_impact"
        ]
        if args.output:
            csv_writer = csv.DictWriter(out_f, fieldnames=output_fieldnames)
            csv_writer.writeheader()
        elif not consolidator:
            csv_writer = csv.DictWriter(sys.stdout, fieldnames=output_fieldnames)
            csv_writer.writeheader()

    def _process(row: dict):
        text = extract_text(row)
        if not text:
            return None
        return row, categorize_text(client, text)

    worker_count = args.workers if args.workers is not None else settings.cli_workers
    if not args.resume_consolidate_from:
        logging.info("Traitement en parallèle: %d workers", worker_count)
        with ThreadPoolExecutor(max_workers=worker_count) as ex, tqdm(total=len(rows), desc="Traitement", unit="ligne") as pbar:
            for res in ex.map(_process, rows):
                if res is None:
                    pbar.update(1)
                    continue
                row, result = res
                cat, subs = taxo.align(result.category, result.subcategories)

                output_row = row.copy()
                output_row.update({
                    "category": cat,
                    "subcategories": json.dumps(subs) if args.format == "csv" else subs,
                    "sentiment": result.sentiment,
                    "emotional_tone": result.emotional_tone,
                    "summary": result.summary,
                    "estimated_impact": result.estimated_impact,
                })

                if consolidator:
                    buffered.append({
                        "original_row": row,
                        "category": cat,
                        "subcategories": subs,
                        "sentiment": result.sentiment,
                        "emotional_tone": result.emotional_tone,
                        "summary": result.summary,
                        "estimated_impact": result.estimated_impact,
                    })
                    all_cats.add(cat)
                    for s in subs:
                        all_subs.add(s)
                    cat_freq[cat] += 1
                else:
                    if args.format == "csv":
                        if csv_writer:
                            csv_writer.writerow(output_row)
                    else:  # jsonl
                        payload = {
                            "ticket": row.get("ticket_id") or row.get("id") or count + 1,
                            "category": cat,
                            "subcategories": subs,
                            "sentiment": result.sentiment,
                            "emotional_tone": result.emotional_tone,
                            "summary": result.summary,
                            "estimated_impact": result.estimated_impact,
                        }
                        data = orjson.dumps(payload)
                        if out_f:
                            out_f.write(data + b"\n")
                        else:
                            sys.stdout.buffer.write(data + b"\n")

                    count += 1
                    pbar.update(1)

        # Sauvegarder l'intermédiaire si demandé
        if consolidator and args.save_intermediate:
            logging.info("Sauvegarde de l'intermédiaire: %s", args.save_intermediate)
            with args.save_intermediate.open("wb") as f:
                for item in buffered:
                    f.write(orjson.dumps(item) + b"\n")
    else:
        # Mode reprise: charger le JSONL intermédiaire et configurer consolidation
        if not args.resume_consolidate_from.exists():
            logging.error("Fichier intermédiaire introuvable: %s", args.resume_consolidate_from)
            return 2
        logging.info("Reprise de la consolidation depuis %s", args.resume_consolidate_from)
        consolidator = LLMTaxonomyConsolidator(client)
        buffered = []
        all_cats.clear(); all_subs.clear()
        cat_freq.clear()
        with args.resume_consolidate_from.open("rb") as f:
            for line in f:
                if not line.strip():
                    continue
                obj = orjson.loads(line)
                if not isinstance(obj, dict) or "category" not in obj or "subcategories" not in obj:
                    logging.warning("Entrée intermédiaire invalide ignorée")
                    continue
                buffered.append(obj)
                c = obj["category"]
                all_cats.add(c)
                cat_freq[c] += 1
                for s in obj["subcategories"]:
                    all_subs.add(s)

    # Consolidation
    if consolidator and buffered:
        batch_size = args.consolidation_batch_size if args.consolidation_batch_size is not None else settings.consolidation_batch_size
        # Heuristique de pré-fusion des catégories
        from .taxonomy import heuristic_category_mapping
        pre_map = heuristic_category_mapping(
            all_cats,
            cat_freq,
            synonyms=settings.taxonomy_synonyms,
            max_categories=(args.max_categories if args.max_categories is not None else settings.max_categories),
            threshold=(args.merge_threshold if args.merge_threshold is not None else settings.merge_threshold),
        )
        if pre_map:
            logging.info("Heuristic pre-merge: categories=%d -> %d (threshold=%.2f, max=%s)",
                         len(all_cats), len(set(pre_map.values())),
                         (args.merge_threshold if args.merge_threshold is not None else settings.merge_threshold),
                         (args.max_categories if args.max_categories is not None else settings.max_categories))
        # Calcul du nombre de lots pour la barre de progression
        def _nb_batches(n: int, b: int) -> int:
            return (n + max(1, b) - 1) // max(1, b)
        total_batches = _nb_batches(len(all_cats), batch_size) + _nb_batches(len(all_subs), batch_size)
        if total_batches <= 0:
            total_batches = 1
        logging.info("Consolidation: categories=%d, subcategories=%d, batch_size=%d", len(all_cats), len(all_subs), batch_size)
        with tqdm(total=total_batches, desc="Consolidation LLM", unit="lot") as pbar:
            cat_map, sub_map = consolidator.consolidate_batched(
                all_cats,
                all_subs,
                batch_size=batch_size,
                on_progress=pbar.update,
                seed_category_mapping=pre_map,
            )
        logging.info("Consolidation mappings: category_mapping=%d, subcategory_mapping=%d", len(cat_map), len(sub_map))

        # Passes supplémentaires pour harmoniser entre lots
        rounds = args.consolidation_rounds if args.consolidation_rounds is not None else settings.consolidation_rounds
        for r in range(2, max(2, rounds) + 1):
            prev_cats = len(set(cat_map.values()))
            prev_subs = len(set(sub_map.values()))
            cats2 = set(cat_map.values())
            subs2 = set(sub_map.values())
            total_batches2 = _nb_batches(len(cats2), batch_size) + _nb_batches(len(subs2), batch_size)
            if total_batches2 <= 0:
                break
            with tqdm(total=total_batches2, desc=f"Consolidation LLM (round {r})", unit="lot") as pbar2:
                cat_map2, sub_map2 = consolidator.consolidate_batched(cats2, subs2, batch_size=batch_size, on_progress=pbar2.update)
            cat_map = {k: cat_map2.get(v, v) for k, v in cat_map.items()}
            sub_map = {k: sub_map2.get(v, v) for k, v in sub_map.items()}
            if len(set(cat_map.values())) == prev_cats and len(set(sub_map.values())) == prev_subs:
                break
        logging.info("Consolidation unique counts after mapping: categories=%d -> %d, subcategories=%d -> %d",
                    len(all_cats), len(set(cat_map.values())), len(all_subs), len(set(sub_map.values())))

        # Initialiser csv_writer pour consolidation si nécessaire
        if args.format == "csv" and not csv_writer:
            if not input_fieldnames and buffered:
                # Construire en-têtes depuis original_row (mode reprise)
                base = list((buffered[0].get("original_row") or {}).keys())
                seen = set(base)
                extras: list[str] = []
                for it in buffered:
                    orow = it.get("original_row") or {}
                    for k in orow.keys():
                        if k not in seen:
                            seen.add(k)
                            extras.append(k)
                input_fieldnames = base + extras
            output_fieldnames = input_fieldnames + [
                "category", "subcategories", "sentiment",
                "emotional_tone", "summary", "estimated_impact"
            ]
            csv_writer = csv.DictWriter(sys.stdout, fieldnames=output_fieldnames)
            csv_writer.writeheader()

        final_cats: set[str] = set()
        final_subs: set[str] = set()
        for item in tqdm(buffered, desc="Consolidation", unit="ligne"):
            row = item["original_row"]
            cat = cat_map.get(item["category"], item["category"])
            subs = [sub_map.get(s, s) for s in item["subcategories"]]
            final_cats.add(cat)
            for s in subs:
                final_subs.add(s)

            if args.format == "csv":
                output_row = row.copy()
                output_row.update({
                    "category": cat,
                    "subcategories": json.dumps(subs),
                    "sentiment": item["sentiment"],
                    "emotional_tone": item["emotional_tone"],
                    "summary": item["summary"],
                    "estimated_impact": item["estimated_impact"],
                })
                if csv_writer:
                    csv_writer.writerow(output_row)
            else:  # jsonl
                payload = {
                    "ticket": row.get("ticket_id") or row.get("id") or "N/A",
                    "category": cat,
                    "subcategories": subs,
                    "sentiment": item["sentiment"],
                    "emotional_tone": item["emotional_tone"],
                    "summary": item["summary"],
                    "estimated_impact": item["estimated_impact"],
                }
                data = orjson.dumps(payload)
                if out_f:
                    out_f.write(data + b"\n")
                else:
                    sys.stdout.buffer.write(data + b"\n")

    processed_total = count if count > 0 else (len(buffered) if consolidator else 0)
    if consolidator and buffered:
        logging.info("Final consolidated unique counts: categories=%d, subcategories=%d", len(final_cats), len(final_subs))
    else:
        logging.info(
            "Processed %d feedback(s). Categories=%d, Subcategories(total)=%d",
            processed_total,
            len(taxo.categories),
            sum(len(v) for v in taxo.subs_by_cat.values()),
        )

    if out_f:
        out_f.close()
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
